import sys
import calendar
import datetime
import sys

import pandas as pd
import numpy as np
import openpyxl

from FoodBankManager import FBM

donor_catagories = ["Grocery", "Org/Corp", "Individual", "Church", "Purchased", "Senior program"]

def FindLastMonthsDates():
	"""
	Finds the first and last days of last month
	:return: (first_of_month, last_of_month)
	:rtype: tuple(datetime.date, datetime.date)
	"""
	now = datetime.date.today()
	last_of_month = now.replace(day=1) - datetime.timedelta(days=1)
	first_of_month = last_of_month.replace(day=1)

	return first_of_month, last_of_month

def RunMonthlyReport(FBMInst,month=None, year=None):
	if month is not None and year is not None:
		start = datetime.date(year, month, 1)
		end = datetime.date(year, month, calendar.monthrange(year, month)[1])
	else:
		start, end = FindLastMonthsDates()

	data = FBMInst.GetFoodDonations(start, end)

	data[["Weight (lbs)"]] = data[["Weight (lbs)"]].astype("float")

	return data

def PivotInventoryTable(df):
	df[["Weight (lbs)"]] = df[["Weight (lbs)"]].astype("float")
	df = pd.pivot_table(df, index=["DonorCategory"], values=["Weight (lbs)"], aggfunc=[np.sum])
	return df

def MonthlyGuestData(FBMInst,month=None, year=None):
	if month is not None and year is not None:
		start = datetime.date(year, month, 1)
		end = datetime.date(year, month, calendar.monthrange(year, month)[1])
	else:
		start, end = FindLastMonthsDates()

	data = FBMInst.GetGuestData(start, end)

	data[["Tracking Result"]] = data[["Tracking Result"]].astype("int")

	return data


def WriteSummaryData(q, ws, origin=(1,1),  month=None, year=None):
	if month is not None and year is not None:
		start = datetime.date(year, month, 1)
		end = datetime.date(year, month, calendar.monthrange(year, month)[1])
	else:
		start, end = FindLastMonthsDates()

	donation_data = RunMonthlyReport(q, month=start.month, year=start.year)
	user_data = MonthlyGuestData(q, month=start.month, year=start.year)

	ws.column_dimensions[ws.cell(row=origin[0], column=origin[1]).column].width = 14

	for cell_row in range(origin[0], origin[0] + len(donor_catagories) + 10 + 1):
		ws.cell(row=cell_row, column=origin[1]).style = "Input"

	ws.cell(row=origin[0], column=origin[1]).value = "{}".format(start.strftime("%B %Y"))
	ws.cell(row=origin[0], column=origin[1]).style = "Headline 4"
	ws.cell(row=origin[0], column=origin[1]).alignment = openpyxl.styles.Alignment(horizontal='center')
	for i, item in enumerate(donor_catagories):
		ws.cell(row=origin[0] + i + 1, column=origin[1]).value = donation_data[donation_data["DonorCategory"] == item].sum()["Weight (lbs)"]
	ws.cell(row=origin[0] + len(donor_catagories) + 1, column=origin[1]).value = "=SUM({}:{})".format(ws.cell(row=origin[0] + 1, column=origin[1]).coordinate, ws.cell(row=origin[0] + len(donor_catagories), column=origin[1]).coordinate)
	ws.cell(row=origin[0] + len(donor_catagories) + 1, column=origin[1]).style = "Calculation"
	ws.cell(row=origin[0] + len(donor_catagories) + 2, column=origin[1]).value = donation_data[donation_data["DonorCategory"] == "Waste"].sum()["Weight (lbs)"]
	ws.cell(row=origin[0] + len(donor_catagories) + 3, column=origin[1]).value = "={}-{}".format(ws.cell(row=origin[0] + len(donor_catagories) + 1, column=origin[1]).coordinate, ws.cell(row=origin[0] + len(donor_catagories) + 2, column=origin[1]).coordinate)
	ws.cell(row=origin[0] + len(donor_catagories) + 3, column=origin[1]).style = "Calculation"
	ws.cell(row=origin[0] + len(donor_catagories) + 4, column=origin[1]).style = "Normal"
	ws.cell(row=origin[0] + len(donor_catagories) + 5, column=origin[1]).value = len(user_data.index)

	ws.cell(row=origin[0] + len(donor_catagories) + 7, column=origin[1]).value = user_data["Tracking Result"].sum()
	ws.cell(row=origin[0] + len(donor_catagories) + 8, column=origin[1]).value = "={}*40".format(ws.cell(row=origin[0] + len(donor_catagories) + 5, column=origin[1]).coordinate)
	ws.cell(row=origin[0] + len(donor_catagories) + 9, column=origin[1]).style = "Normal"
	ws.cell(row=origin[0] + len(donor_catagories) + 10, column=origin[1]).value = "=IF(ISTEXT({0:}), {2:}-{1:}, {2:}-{1:}+{0:})".format(ws.cell(row=origin[0] + len(donor_catagories) + 10, column=origin[1]-1).coordinate, ws.cell(row=origin[0] + len(donor_catagories) + 8, column=origin[1]).coordinate, ws.cell(row=origin[0] + len(donor_catagories) + 3, column=origin[1]).coordinate)


def WriteSummaryLabel(ws, origin=(1,1)):
	ws.column_dimensions[ws.cell(row=origin[0], column=origin[1]).column].width = 22
	for i, item in enumerate(donor_catagories):
		ws.cell(row=origin[0] + i + 1, column=origin[1]).value = "{} (lbs)".format(item)
	ws.cell(row=origin[0] + len(donor_catagories) + 1, column=origin[1]).value = "Total Food Income (lbs)"
	ws.cell(row=origin[0] + len(donor_catagories) + 2, column=origin[1]).value = "Waste (lbs)"
	ws.cell(row=origin[0] + len(donor_catagories) + 3, column=origin[1]).value = "Total Collected (lbs)"

	ws.cell(row=origin[0] + len(donor_catagories) + 5, column=origin[1]).value = "Number of Clients"
	ws.cell(row=origin[0] + len(donor_catagories) + 6, column=origin[1]).value = "New Clients"
	ws.cell(row=origin[0] + len(donor_catagories) + 7, column=origin[1]).value = "Total Impact"
	ws.cell(row=origin[0] + len(donor_catagories) + 8, column=origin[1]).value = "Food Distributed (lbs)"

	ws.cell(row=origin[0] + len(donor_catagories) + 10, column=origin[1]).value = "Ending Inventory (lbs)"

	for cell_row in range(origin[0], origin[0] + len(donor_catagories) + 10 + 1):
		ws.cell(row=cell_row, column=origin[1]).style = "Headline 4"

def WriteExcelSheet(name, month=None, year=None):
	if month is not None and year is not None:
		start = datetime.date(year, month, 1)
		end = datetime.date(year, month, calendar.monthrange(year, month)[1])
	else:
		start, end = FindLastMonthsDates()

	q = FBM("mcfb.soxbox.co")

	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "12 Month Overview"

	WriteSummaryLabel(ws, origin=(2, 1))
	ym_start = 12 * year + month - 1
	for i, ym in enumerate(range(ym_start - 11, ym_start + 1)):
		y, m = divmod(ym, 12)
		WriteSummaryData(q, ws, origin=(2, i+2), month=m+1, year=y)

	ws.merge_cells('B1:M1')
	for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
		ws['{}1'.format(col)].style = "Headline 1"
	ws['B1'] = "12 Month Overview"
	ws['B1'].alignment = openpyxl.styles.Alignment(horizontal='center')

	ws['N1'] = "Last Year's Performance"
	ws['N1'].style = "Headline 1"
	WriteSummaryData(q, ws, origin=(2, 14), month=month, year=year-1)
	ws.column_dimensions['N'].width = 30
	ws['N18'].style = "Normal"
	ws['N18'] = ""

	ws.freeze_panes = "B3"

	wb.save("{}.xlsx".format(name))
	return "{}.xlsx".format(name)

if __name__ == '__main__':
	pd.set_option('display.expand_frame_repr', False)
	if len(sys.argv) < 3:
		print "Run with \"<month number (1-12)> <Year (4 digit)>\""
	print WriteExcelSheet("Report {}-{}".format(sys.argv[1], sys.argv[2]), month=int(sys.argv[1]), year=int(sys.argv[2]))
